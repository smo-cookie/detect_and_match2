import os
import json
import re
import openai
from pymongo import MongoClient
from docx import Document
from openpyxl import load_workbook

# open api 키 설정해야함
openai.api_key = ""

# mongo db 설정 후 연결
client = MongoClient("")
db = client["personal_info_db"]  # 사용자 개인정보(ex> 비밀번호)
collection = db["detected_info"] # 탐지된 개인정보(ex> 연락처, 주민등록번호 등등...)

# 정규표현식 패턴 -> 지정해 놓은 양식만 탐지(ex> 010-xxxx-xxxx 이런 형식만 탐지가능)
patterns = {
    "주민등록번호": r"\b\d{6}-\d{7}\b",
    # "주소": r"\b[가-힣]+시 [가-힣]+구 [가-힣]+동\b",
    "연락처": r"\b010-\d{4}-\d{4}\b",
    "생년월일": r"\b\d{4}[-/]\d{2}[-/]\d{2}\b",
    "계좌번호": r"\b\d{2,4}-\d{2,4}-\d{2,4}\b",
    "여권번호": r"\b[A-Z]{1}\d{8}\b",
    "이메일": r"\b[A-Za-z0-9._%+-]+@(?:[A-Za-z0-9-]+\.)+[A-Za-z]{2,}\b",
    "카드번호": r"\b\d{4}-\d{4}-\d{4}-\d{4}\b",
    # "성명": r"\b[가-힣]{2,3}\b"
}


def extract_text_from_word(file_path):
    # word에서 추출
    document = Document(file_path)
    return "\n".join([paragraph.text for paragraph in document.paragraphs])


def extract_text_from_excel(file_path):
    # excel에서 추출
    workbook = load_workbook(file_path)
    text = ""
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only=True):
            text += " ".join([str(cell) if cell else "" for cell in row]) + "\n"
    return text


def detect_pii_with_regex(content):
    # 정규표현식 탐지
    results = {}
    for key, pattern in patterns.items():
        matches = re.findall(pattern, content)
        if matches:
            results[key] = matches
    return results


def merge_results(regex_results, chatgpt_results):
    # 결과 통합 부분 -> 정규표현식으로 먼저 탐지 후 챗 지피티가 교차 검증
    merged_results = {}
    for key in patterns.keys():
        regex_data = set(regex_results.get(key, [])) 
        chatgpt_data = set(chatgpt_results.get(key, []))
        merged_results[key] = list(regex_data | chatgpt_data)
    return merged_results


def detect_sensitive_info_with_chatgpt(content, additional_info): # llm 요청 시 좀 더 세밀한 질문으로 변경 필요, 필요시 영어로 요청
    """ChatGPT API를 사용하여 추가 탐지"""
    prompt = f"""
    다음 텍스트에서 개인정보와 추가 요청된 정보를 탐지해주세요:
    - 개인정보에는 연락처, 이메일, 주민등록번호, 주소, 계좌번호 등 개인을 특정할 수 있는 정보가 포함됩니다.
    - 추가 요청 정보: {additional_info}
    반환 형식:
    {{
        "개인정보": {{
            "연락처": ["01012345678", ...],
            "이메일": ["example@domain.com", ...],
            ...
        }},
        "추가 탐지 정보": {{
            "추가 요청 정보": ["Project Alpha", ...],
        }}
    }}
    텍스트:
    {content}
    """
    response = openai.ChatCompletion.create(
        model="gpt-4",
        messages=[{"role": "user", "content": prompt}]
    )
    try:
        return json.loads(response['choices'][0]['message']['content'])
    except json.JSONDecodeError:
        return {"error": "Invalid JSON from ChatGPT"}


def save_to_mongodb(file_name, file_content, final_results, additional_results):
    # 세팅된 몽고 디비에 저장
    document = {
        "file_name": file_name, # 파일명
        "file_content": file_content, # 파일 전체 내용
        "detected_info": final_results, # 통합된 개인정보 탐지 결과
        "chatgpt_plus_results": additional_results # 추가로 탐지된 결과
    }
    collection.insert_one(document)


def main(file_path, file_type, additional_info):
    # 파일 형식 검증
    if file_type == "word" and not file_path.endswith(".docx"):
        print("지원하지 않는 Word 파일 형식입니다. .docx 파일만 지원됩니다.")
        return
    elif file_type == "excel" and not file_path.endswith(".xlsx"):
        print("지원하지 않는 Excel 파일 형식입니다. .xlsx 파일만 지원됩니다.")
        return

    # 문서 내용 추출
    if file_type == "word":
        content = extract_text_from_word(file_path)
    elif file_type == "excel":
        content = extract_text_from_excel(file_path)
    else:
        print("지원하지 않는 파일 형식입니다.")
        return

    # 정규표현식 탐지
    regex_results = detect_pii_with_regex(content)

    # 챗지피티를 통한 추가 탐지
    chatgpt_response = detect_sensitive_info_with_chatgpt(content, additional_info)
    if "error" in chatgpt_response:
        print("ChatGPT 탐지 중 오류 발생:", chatgpt_response["error"])
        return

    # 챗지피티 탐지 결과
    chatgpt_results = chatgpt_response.get("개인정보", {})
    additional_results = chatgpt_response.get("추가 탐지 정보", {})

    # 정규표현식 결과와 챗지피티 결과 통합 -> 사용자 gui 창에 내보낼 것
    final_results = merge_results(regex_results, chatgpt_results)

    # 결과 저장
    save_to_mongodb(file_path, content, final_results, additional_results)

    # 디버깅
    print(json.dumps({
        "final_results": final_results,
        "chatgpt_plus_results": additional_results
    }, ensure_ascii=False, indent=4))


if __name__ == "__main__":
    import sys
    file_path = sys.argv[1]  # 파일 경로
    file_type = sys.argv[2]  # 파일 타입
    additional_info = sys.argv[3]  # 추가 탐지 요청 정보 -> 사용자로부터 선택적으로 입력받음
    main(file_path, file_type, additional_info)
