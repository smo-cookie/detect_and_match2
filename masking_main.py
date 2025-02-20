import os
import json
import re
import openai
import zipfile
from pymongo import MongoClient
from docx import Document
from openpyxl import load_workbook
from tempfile import TemporaryDirectory
from lxml import etree

openai.api_key = ""

MONGO_URI = "mongodb+srv://smocookie:smocookie@cluster0.btwrt.mongodb.net/?retryWrites=true&w=majority"
client = MongoClient(MONGO_URI)
db = client["personal_info_db"]
detected_info_collection = db["detected_info"]
file_metadata_collection = db["file_metadata"]
additional_info_collection = db["additional_info"]

patterns = {
    "주민등록번호": r"\b\d{6}-\d{7}\b",
    "연락처": r"\b010-\d{4}-\d{4}\b",
    "생년월일": r"\b\d{4}[-/]\d{2}[-/]\d{2}\b",
    "계좌번호": r"\b\d{2,4}-\d{2,4}-\d{2,4}\b",
    "여권번호": r"\b[A-Z]{1}\d{8}\b",
    "이메일": r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b",
    "카드번호": r"\b\d{4}-\d{4}-\d{4}-\d{4}\b"
}


def detect_pii_with_regex(content):
    results = {}
    for key, pattern in patterns.items():
        matches = re.findall(pattern, content)
        if matches:
            results[key] = list(set(matches))  # 중복 제거
    return results


def detect_sensitive_info_with_chatgpt(content, additional_info):
    prompt = f"""
    다음 텍스트에서 개인정보(이름 및 주소)와 추가 요청된 정보를 탐지해주세요:
    - 개인정보에는 연락처, 이메일, 주민등록번호, 주소, 계좌번호 등 개인을 특정할 수 있는 정보가 포함됩니다.
    - 추가 요청 정보: {additional_info}
    반환 형식(JSON):
    {{
        "개인정보": {{
            "이름": ["홍길동", "김철수"],
            "주소": ["서울시 강남구 역삼동"]
        }},
        "추가 탐지 정보": {{
            "추가 요청 정보": ["Project Alpha", "XYZ Corporation"]
        }}
    }}
    텍스트:
    {content}
    """
    response = openai.ChatCompletion.create(
        model="gpt-4",
        messages=[{"role": "user", "content": prompt}]
    )

    try:
        return json.loads(response['choices'][0]['message']['content'])
    except json.JSONDecodeError:
        return {"error": "Invalid JSON from ChatGPT"}


# def save_to_mongodb(file_name, detected_info, additional_results):
#     document = {
#         "file_name": file_name,
#         "detected_info": detected_info,
#         "chatgpt_plus_results": additional_results
#     }
#     collection.insert_one(document)

def save_to_mongodb(file_name, detected_info, additional_results):
    file_metadata_collection.insert_one({"file_name": file_name})
    detected_info_collection.insert_one({"file_name": file_name, "detected_info": detected_info})
    additional_info_collection.insert_one({"file_name": file_name, "additional_info": additional_results})


def get_masking_data_from_mongodb(file_name):
    masking_data = set()
    detected_info = detected_info_collection.find_one({"file_name": file_name})
    if detected_info and "detected_info" in detected_info:
        for values in detected_info["detected_info"].values():
            masking_data.update(values)

    # 사용자가 추가한 마스킹 정보 가져오기
    additional_info = additional_info_collection.find_one({"file_name": file_name})
    if additional_info and "additional_info" in additional_info:
        masking_data.update(additional_info["additional_info"])

    return masking_data

# 마스킹 - 정규표현식
def apply_masking(content, masking_data):
    for item in masking_data:
        content = content.replace(item, "****")
    return content

# word 문서 텍스트 추출
def extract_text_from_word(file_path):
    document = Document(file_path)
    return "\n".join([paragraph.text for paragraph in document.paragraphs])

# excel 문서 텍스트 추출
def extract_text_from_excel(file_path):
    workbook = load_workbook(file_path)
    text = ""
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only=True):
            text += " ".join([str(cell) if cell else "" for cell in row]) + "\n"
    return text

# xml 기반 마스킹 적용
def process_xml_file(xml_path, masking_data):
    parser = etree.XMLParser(remove_blank_text=True)
    with open(xml_path, 'rb') as file:
        xml_tree = etree.parse(file, parser)

    for element in xml_tree.xpath("//w:t", namespaces={"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}):
        if element.text:
            element.text = apply_masking(element.text, masking_data)

    with open(xml_path, 'wb') as file:
        file.write(etree.tostring(xml_tree, pretty_print=True))

# 워드 파일 마스킹
def mask_sensitive_data_with_images(file_path):
    masking_data = get_masking_data_from_mongodb(file_path) # 몽고db에서 가져옴

    with TemporaryDirectory() as temp_dir:
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)

        document_xml_path = os.path.join(temp_dir, "word", "document.xml")
        if os.path.exists(document_xml_path):
            process_xml_file(document_xml_path, masking_data)

        new_file_path = file_path.replace(".docx", "(masked).docx")
        with zipfile.ZipFile(new_file_path, 'w') as zip_out:
            for foldername, subfolders, filenames in os.walk(temp_dir):
                for filename in filenames:
                    file_path = os.path.join(foldername, filename)
                    arcname = os.path.relpath(file_path, temp_dir)
                    zip_out.write(file_path, arcname)

    return new_file_path

# 메인 실행
def main(file_path, file_type, additional_info_json):
    additional_info = json.loads(additional_info_json) if additional_info_json else []
    print(f"📂 Processing file: {file_path}")
    print(f"📄 File type: {file_type}")
    print(f"🔍 Additional masking info: {additional_info}")
    if file_type == "word":
        content = extract_text_from_word(file_path)
    elif file_type == "excel":
        content = extract_text_from_excel(file_path)
    else:
        print("지원하지 않는 파일 형식입니다.")
        return

    regex_results = detect_pii_with_regex(content)
    chatgpt_response = detect_sensitive_info_with_chatgpt(content, additional_info)

    if "error" in chatgpt_response:
        print("ChatGPT 탐지 중 오류 발생:", chatgpt_response["error"])
        return

    chatgpt_results = chatgpt_response.get("개인정보", {})
    additional_results = chatgpt_response.get("추가 탐지 정보", {})

    final_results = {**regex_results, **chatgpt_results} # 키, 밸류 값 모두 넣기

    save_to_mongodb(file_path, final_results, additional_results)

    masked_file = mask_sensitive_data_with_images(file_path)

    print(f"마스킹된 파일이 저장되었습니다: {masked_file}")

# 프로그램 시작 시 실행
if __name__ == "__main__":
    import sys
    file_path = sys.argv[1]
    file_type = sys.argv[2]
    additional_info_json = sys.argv[3] if len(sys.argv) > 3 else "[]"
    
    # json 문자열을 리스트로 변환
    additional_info = json.loads(additional_info_json)
    
    main(file_path, file_type, additional_info)
